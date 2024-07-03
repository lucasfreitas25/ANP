from urllib.request import urlretrieve 
from zipfile import ZipFile
import pandas as pd 
import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from ajustar_planilha import ajustar_bordas, ajustar_colunas

url = 'https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos/arquivos/pb-da-biodiesel.zip'
arquivo = 'pb-da-biodiesel.zip'

urlretrieve(url, arquivo)

with ZipFile(arquivo, "r") as zip:
    zip.extractall('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados em csv')

#CAPACIDADE/TANCAGEM    
df_capacidade = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados em csv\\Biodiesel_DadosAbertos_CSV_Capacidade.csv', dtype={'CNPJ':str})
#APLICA MUDANÇA NA DATA 
df_capacidade.rename(columns={'Mês/Ano': 'Data'}, inplace=True)
df_capacidade['Data'] = pd.to_datetime(df_capacidade['Data'], format='%m/%Y', errors='coerce')
df_capacidade['Data'] = df_capacidade['Data'].dt.strftime('%d/%m/%Y') #VOLTA PARA STRING NO FORMATO DESEJADO

df_tanca = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados em csv\\Biodiesel_DadosAbertos_CSV_Tancagem.csv', dtype={'CNPJ':str})
df_tanca.rename(columns={'Mês/Ano': 'Data'}, inplace=True)
df_tanca['Data'] = pd.to_datetime(df_tanca['Data'], format='%m/%Y', errors='coerce')
df_tanca['Data'] = df_tanca['Data'].dt.strftime('%d/%m/%Y') 
df_capacidade = pd.merge(df_capacidade, df_tanca, on=['Data', 'Razão Social', 'CNPJ', 'Estado', 'Região', 'Município'], how='inner')

df_capacidade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Capacidade.xlsx', index=False)
df_capacidade.to_html('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Capacidade.html', index=False)
df_capacidade.to_string('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Capacidade.txt', index=False)

#MATERIA PRIMA
df_matprima = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados em csv\\Biodiesel_DadosAbertos_CSV_MatériaPrima.csv')
df_matprima['Quantidade (m³)'] = df_matprima['Quantidade (m³)'].str.replace(',', '.').astype(float)
df_matprima['Produto'] = df_matprima['Produto'].replace('ÓLEO DE ALGODÃO (GOSSYPIUM HIRSUT)', 'ÓLEO DE ALGODÃO').replace('ÓLEO DE COLZA/CANOLA (BRESSICA CAMPESTRIS)', \
    'ÓLEO DE COLZA/CANOLA').replace('ÓLEO DE PALMA/DENDÊ (ELAEIS GUINEENSIS OU ELAEIS O', 'ÓLEO DE PALMA/DENDÊ').replace('ÓLEO DE SOJA (GLYCINE MAX)', 'ÓLEO DE SOJA')
df_matprima.rename(columns={'Mês/Ano': 'Data'}, inplace=True)
df_matprima['Data'] = pd.to_datetime(df_matprima['Data'], format='%m/%Y', errors='coerce')
df_matprima['Data'] = df_matprima['Data'].dt.strftime('%d/%m/%Y') 

df_matprima.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Materia Prima.xlsx', index=False)
df_matprima.to_html('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Materia Prima.html', index=False)
df_matprima.to_string('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Materia Prima.txt', index=False)

#PRODUÇÃO
df_prod = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados em csv\\Biodiesel_DadosAbertos_CSV_Produç╞o.csv')
df_prod['Produção de Biodiesel'] = df_prod['Produção de Biodiesel'].str.replace(',', '.').astype(float)
df_prod.rename(columns={'Mês/Ano': 'Data'}, inplace=True)
df_prod['Data'] = pd.to_datetime(df_prod['Data'], format='%m/%Y', errors='coerce')
df_prod['Data'] = df_prod['Data'].dt.strftime('%d/%m/%Y') 

df_prod.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Produção.xlsx', index=False)
df_prod.to_html('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Producao.html', index=False)
df_prod.to_string('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Producao.txt', index=False)


#VENDAS
df_vendas = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados em csv\\Biodiesel_DadosAbertos_CSV_Vendas.csv')
df_vendas['Vendas de Biodiesel'] = df_vendas['Vendas de Biodiesel'].str.replace(',', '.').astype(float)
df_vendas.rename(columns={'Mês/Ano': 'Data'}, inplace=True)
df_vendas['Data'] = pd.to_datetime(df_vendas['Data'], format='%m/%Y', errors='coerce')
df_vendas['Data'] = df_vendas['Data'].dt.strftime('%d/%m/%Y') 

df_vendas.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Vendas.xlsx', index=False)
df_vendas.to_html('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Vendas.html', index=False)
df_vendas.to_string('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\CHATBOT\\Banco de dados Bot\\Dados Biodiesel Vendas.html', index=False)

planilha_principal = openpyxl.Workbook()

wb_capacidade = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Capacidade.xlsx')  
wb_matprima = openpyxl.load_workbook("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Materia Prima.xlsx")  
wb_prod = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Produção.xlsx')
wb_vendas = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\Dados Biodiesel Vendas.xlsx')


aba_capacidade = planilha_principal.create_sheet("CAPACIDADE")
aba_matprima = planilha_principal.create_sheet("MATÉRIA PRIMA")
aba_prod = planilha_principal.create_sheet("PRODUÇÃO")
aba_vendas = planilha_principal.create_sheet("VENDAS")


for linha in wb_capacidade.active.iter_rows(values_only=True):
    aba_capacidade.append(linha)

for linha in wb_matprima.active.iter_rows(values_only=True):
    aba_matprima.append(linha)
    
for linha in wb_prod.active.iter_rows(values_only=True):
    aba_prod.append(linha)
    
for linha in wb_vendas.active.iter_rows(values_only=True):
    aba_vendas.append(linha)

for aba in planilha_principal.sheetnames:
    if aba not in ["CAPACIDADE", "MATÉRIA PRIMA", "PRODUÇÃO", "VENDAS"]:
        del planilha_principal[aba]

lista_abas = [aba_capacidade, aba_matprima, aba_prod, aba_vendas]
for abas in lista_abas:
    ajustar_colunas(abas)
    
############################    
worksheet = planilha_principal.active

ajustar_bordas(planilha_principal)

planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\BIODIESEL\\BIODIESEL ANP.xlsx")
