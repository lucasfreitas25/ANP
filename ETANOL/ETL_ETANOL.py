from urllib.request import urlretrieve 
from zipfile import ZipFile
import pandas as pd 
import openpyxl
from ajustar_planilha import ajustar_bordas, ajustar_colunas
from Google import Create_Service
from Drive import add_arquivos_a_pasta

#BAIXA O ARQUIVO E EXTRAI ELE PARA UMA PASTA
url = 'https://www.gov.br/anp/pt-br/assuntos/producao-e-fornecimento-de-biocombustiveis/etanol/arquivos-etanol/pb-da-etanol.zip'
arquivo = 'pb-da-etanol.zip'

urlretrieve(url, arquivo)

with ZipFile(arquivo, "r") as zip:
    zip.extractall('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados em csv')
    
def etl_no_df(df):
    df.rename(columns={'Mês/Ano': 'Data'}, inplace=True)
    df['Data'] = pd.to_datetime(df['Data'], format='%m/%Y', errors='coerce')
    df['Data'] = df['Data'].dt.strftime('%d/%m/%Y') #VOLTA PARA STRING NO FORMATO DESEJADO   
#FAZ LIMPEZA E MUDANÇAS ESTRUTURAIS
df_capacidade = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados em csv\\Etanol_DadosAbertos_CSV_Capacidade.csv', dtype={'CNPJ': str})
etl_no_df(df_capacidade)
df_capacidade.fillna(value=0, inplace=True)

df_capacidade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados Etanol Capacidade.xlsx', index=False)
df_capacidade.to_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados Etanol Capacidade.csv', index=False)

df_matprima = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados em csv\\Etanol_DadosAbertos_CSV_MatériaPrima.csv')
df_matprima['Quantidade Processada (t)'] = df_matprima['Quantidade Processada (t)'].str.replace(',', '.').astype(float)
etl_no_df(df_matprima)
df_matprima.fillna(value=0, inplace=True)

df_matprima.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados Etanol Materia Prima.xlsx', index=False)

df_prod = pd.read_csv('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados em csv\\Etanol_DadosAbertos_CSV_Produç╞o.csv')
df_prod.fillna(value=0, inplace=True)
etl_no_df(df_prod)

df_prod.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados Etanol Produção.xlsx', index=False)
print(df_prod)

#JUNTAR TODAS AS PLANILHAS EM UMA
planilha_principal = openpyxl.Workbook()

wb_capacidade = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados Etanol Capacidade.xlsx')  
wb_matprima = openpyxl.load_workbook("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados Etanol Materia Prima.xlsx")  
wb_prod = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\ANP\\ETANOL\\Dados Etanol Produção.xlsx')

aba_capacidade = planilha_principal.create_sheet("CAPACIDADE")
aba_matprima = planilha_principal.create_sheet("MATERIA PRIMA")
aba_prod = planilha_principal.create_sheet("PRODUÇÃO")

for linha in wb_capacidade.active.iter_rows(values_only=True):
    aba_capacidade.append(linha)

for linha in wb_matprima.active.iter_rows(values_only=True):
    aba_matprima.append(linha)
    
for linha in wb_prod.active.iter_rows(values_only=True):
    aba_prod.append(linha)

for aba in planilha_principal.sheetnames:
    if aba not in ["CAPACIDADE", "MATERIA PRIMA", "PRODUÇÃO"]:
        del planilha_principal[aba]

lista_abas = [aba_capacidade, aba_matprima, aba_prod]
for abas in lista_abas:
    ajustar_colunas(abas)
    
############################    
worksheet = planilha_principal.active
ajustar_bordas(planilha_principal)
planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\TABELAS\\TABELAS EM CSV\\ETANOL ANP.xlsx")


'''
CLIENT_SECRET_FILE = 'credencials.json'
API_NAME = 'drive'
API_VERSION = 'v3'
SCOPES = ["https://www.googleapis.com/auth/drive"]

service = Create_Service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)

#PASSA O ARQUIVO PARA O DRIVE
file_id = "1e2yihre6trC07ai7IayhrvLrFP4c57za"
FILE_NAMES = ["ETANOL ANP.xlsx"]
MIME_TYPES = ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]

add_arquivos_a_pasta(FILE_NAMES, MIME_TYPES, service, file_id)
'''
if __name__ == '__main__':
    from sql import executar_sql 
    executar_sql()

